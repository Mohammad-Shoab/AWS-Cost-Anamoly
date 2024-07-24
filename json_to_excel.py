import json
import openpyxl

def json_to_excel(filename, output_filename):
  """
  Reads JSON data from a file and converts it into an Excel sheet with service names as sheets and usage types as columns.

  Args:
      filename: The name of the JSON file to read.
      output_filename: The name of the output Excel file.
  """

  with open(filename, "r") as file:
    try:
      json_data = json.load(file)  # Load JSON data from file
    except FileNotFoundError:
      print(f"Error: File '{filename}' not found.")
      return

  wb = openpyxl.Workbook()

  # Skip the first sheet (avoid creating an empty sheet)
  sheet_index = 0

  # Create sheet with name List of Services
  sheet = wb.create_sheet("List of Services", index=sheet_index)
  sheet_index += 1

  # Add "Services" as the first column header
  sheet.cell(row=1, column=1).value = "Services"

  # Add service names as data in the first column (starting from row 2)
  for row, service_name in enumerate(json_data.keys(), start=2):  # Use enumerate and keys()
    sheet.cell(row=row, column=1).value = service_name

  for service_name, service_data in json_data.items():
    # Create sheet with service name
    sheet = wb.create_sheet(service_name, index=sheet_index)
    sheet_index += 1

    # Extract usage types and costs
    usage_types = service_data["UsageTypes"]
    usage_type_names = list(usage_types.keys())
    usage_costs = list(usage_types.values())

    # Add "UsageTypes" as the first column header
    sheet.cell(row=1, column=1).value = "UsageTypes"

    # Add "Daily-Alert" as the second column header
    sheet.cell(row=1, column=2).value = "Daily-Alert"

    # Add "Daily-Threshold" as the third column header
    sheet.cell(row=1, column=3).value = "Daily-Threshold"

    # Add "Weekly-Alert" as the fourth column header
    sheet.cell(row=1, column=4).value = "Weekly-Alert"

    # Add "Weekly-Threshold" as the fifth column header
    sheet.cell(row=1, column=5).value = "Weekly-Threshold"

    # Add "Monthly-Alert" as the sixth column header
    sheet.cell(row=1, column=6).value = "Monthly-Alert"

    # Add "Monthly-Threshold" as the seventh column header
    sheet.cell(row=1, column=7).value = "Monthly-Threshold"

    # Add usage type names as data in the first column (excluding the first one)
    for row, usage_type in enumerate(usage_type_names, start=2):
      sheet.cell(row=row, column=1).value = usage_type

    # Add usage costs in the second column (starting from the second row)
    #for row, usage_cost in enumerate(usage_costs, start=2):
      #sheet.cell(row=row, column=2).value = usage_cost

    # Add False in the second column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=2).value = "False"

    # Add NA in the third column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=3).value = "NA"

    # Add False in the fourth column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=4).value = "False"

    # Add NA in the fifth column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=5).value = "NA"

    # Add False in the sixth column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=6).value = "False"

    # Add NA in the seventh column (starting from the second row)
    for row, usage_cost in enumerate(usage_costs, start=2):
      sheet.cell(row=row, column=7).value = "NA"

  wb.save(output_filename)  # Save the workbook as an Excel file

# Example usage
json_to_excel("service.json", "service2.xlsx")
